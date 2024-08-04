import win32com.client as client
from openpyxl import load_workbook

import logging
from logging.handlers import TimedRotatingFileHandler


logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        TimedRotatingFileHandler(
            filename="outlook.log", when="s", interval=1, backupCount=5
        )
    ],
)
logger = logging.getLogger("main")
logger.setLevel(logging.DEBUG)

outlook = client.Dispatch("Outlook.Application")
namespace = outlook.GetNameSpace("MAPI")

wb = load_workbook("config.xlsx")
ws = wb.active
for var, val in zip(ws["A"], ws["B"]):
    if var.value == "Accounts":
        accounts = val.value
        accounts = accounts.split(",")
    elif var.value == "Folders":
        folders = val.value
        folders = folders.split(",")
    elif var.value == "Main Sales Account":
        main_sales_account = val.value
    elif var.value == "Offer Page link":
        offer_page_link = val.value
    elif var.value == "PO processing folder":
        po_to_process_folder = val.value
wb.close()
logger.info("Config data loaded")

for account, folder in zip(accounts, folders):

    account = namespace.Folders.Item(account)
    logger.info("Current Account is " + account.Name)
    folder = account.Folders.Item(folder)
    logger.info("Current Folder is " + folder.Name)
    drafts = account.Folders.Item("Drafts")

    mails = [message for message in folder.Items if "PO" in message.Subject]
    logger.info(str(len(mails)) + " mails fetched to be processed")
    for mail in mails:
        subject = mail.Subject
        words = subject.split()
        for word in words:
            if word.startswith("PO"):
                po_number = word
                logger.debug(
                    "Mail for "
                    + po_number
                    + " found in "
                    + account.Name
                    + " under "
                    + folder.Name
                    + " folder."
                )
                break

        try:
            workbook = load_workbook(po_to_process_folder + po_number + ".xlsx")
            worksheet = workbook.active
            logger.debug(po_number + ".xlsx found in the folder")
            To = worksheet["A1"].value
            if To == None:
                To = ""
            Cc = worksheet["B1"].value
            if Cc == None:
                Cc = ""
            Bcc = worksheet["C1"].value
            if Bcc == None:
                Bcc = ""
            Comment = worksheet["D1"].value
            if Comment == None:
                Comment = ""
            Comment = offer_page_link + "\n" + Comment
            workbook.close()

            mail.To = To
            mail.Cc = Cc
            mail.Bcc = Bcc
            if account.Name == "SaleAccount2@xyz.com":
                mail.Bcc = Bcc + ";" + main_sales_account
            mail.Body = mail.Body.format(Comment)
            mail.Send()
            logger.info("Mail " + po_number + " sent.")
        except FileNotFoundError:
            logger.warning(po_number + ".xlsx not found in the folder")
            logger.info("Mail " + po_number + " not sent")

    logger.info("Account " + account.Name + " is processed. Moving to next account")
logger.info("All accounts processed")
